# teststring = '81d110d109d108d107d106d105d104d103d102d-101d48b49d100-d99d98'

# _str = teststring.replace('d48b49d' ,'d48d49d')

# print(_str)

import openpyxl
terminal_num = 198

filePath = r'data\new_data\xlsx\subway_data\distance_route_price.xlsx'
wb = openpyxl.load_workbook(filePath)
sheet = wb['sheet1']

# for i in range(terminal_num * terminal_num):
#     string1 = str(sheet.cell(i + 1, 6).value)
#     string2 = string1.replace('d48b49d' ,'d48d49d')
#     sheet.cell(i + 1, 6).value = string2

for i in range(terminal_num * terminal_num):
    if sheet.cell(i + 1, 1).value == sheet.cell(i + 1, 3).value:
         sheet.cell(i + 1, 6).value = sheet.cell(i + 1, 1).value

print('finished!')
wb.save(filePath)