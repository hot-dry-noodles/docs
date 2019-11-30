# -*- coding: utf-8 -*-

import openpyxl

line1Path = 'line1.xlsx'
line1wb = openpyxl.load_workbook(line1Path)
line1sheet = line1wb['sheet1']

line2Path = 'line2.xlsx'
line2wb = openpyxl.load_workbook(line2Path)
line2sheet = line2wb['sheet1']

line3Path = 'line3.xlsx'
line3wb = openpyxl.load_workbook(line3Path)
line3sheet = line3wb['sheet1']

line4Path = 'line4.xlsx'
line4wb = openpyxl.load_workbook(line4Path)
line4sheet = line4wb['sheet1']

line6Path = 'line6.xlsx'
line6wb = openpyxl.load_workbook(line6Path)
line6sheet = line6wb['sheet1']

line7Path = 'line7.xlsx'
line7wb = openpyxl.load_workbook(line7Path)
line7sheet = line7wb['sheet1']

line8Path = 'line8.xlsx'
line8wb = openpyxl.load_workbook(line8Path)
line8sheet = line8wb['sheet1']

line11Path = 'line11.xlsx'
line11wb = openpyxl.load_workbook(line11Path)
line11sheet = line11wb['sheet1']

line12Path = 'line12.xlsx'
line12wb = openpyxl.load_workbook(line12Path)
line12sheet = line12wb['sheet1']


def createSet(sheet, column_num):
    """ create the set according to the column_num """
    cset = set()
    for i in range(sheet.max_row):
        cset.add(sheet.cell(i + 1, column_num).value)
    return cset


line1set = createSet(line1sheet, 2)
line2set = createSet(line2sheet, 2)
line3set = createSet(line3sheet, 2)
line4set = createSet(line4sheet, 2)
line6set = createSet(line6sheet, 2)
line7set = createSet(line7sheet, 2)
line8set = createSet(line8sheet, 2)
line11set = createSet(line11sheet, 2)
line12set = createSet(line12sheet, 2)

setlist = [line1set,
           line2set,
           line3set,
           line4set,
           line6set,
           line7set,
           line8set,
           line11set,
           line12set]

transfer_set = set()

for i in range(len(setlist)):
    for j in range(len(setlist)):
        if i != j:
            temp_set = setlist[i].intersection(setlist[j])
            transfer_set = transfer_set.union(temp_set)

transfer_list = list(transfer_set)
transfer_list.sort()

transfer_num = len(transfer_list)
row_count = 0

print(transfer_list)

filePath = 'transfer.xlsx'
wb = openpyxl.load_workbook(filePath)
sheet = wb['sheet1']

for i in range(transfer_num):
    sheet.cell(i + 1, 1).value = transfer_list[i]

for i in range(transfer_num):

    for j1 in range(line1sheet.max_row):
        if sheet.cell(i + 1, 1).value == line1sheet.cell(j1 + 1, 2).value:
            sheet.cell(i + 1, 2).value = line1sheet.cell(j1 + 1, 3).value

    for j2 in range(line2sheet.max_row):
        if sheet.cell(i + 1, 1).value == line2sheet.cell(j2 + 1, 2).value:
            sheet.cell(i + 1, 2).value = line2sheet.cell(j2 + 1, 3).value

    for j3 in range(line3sheet.max_row):
        if sheet.cell(i + 1, 1).value == line3sheet.cell(j3 + 1, 2).value:
            sheet.cell(i + 1, 2).value = line3sheet.cell(j3 + 1, 3).value

    for j4 in range(line4sheet.max_row):
        if sheet.cell(i + 1, 1).value == line4sheet.cell(j4 + 1, 2).value:
            sheet.cell(i + 1, 2).value = line4sheet.cell(j4 + 1, 3).value

    for j6 in range(line6sheet.max_row):
        if sheet.cell(i + 1, 1).value == line6sheet.cell(j6 + 1, 2).value:
            sheet.cell(i + 1, 2).value = line6sheet.cell(j6 + 1, 3).value

    for j7 in range(line7sheet.max_row):
        if sheet.cell(i + 1, 1).value == line7sheet.cell(j7 + 1, 2).value:
            sheet.cell(i + 1, 2).value = line7sheet.cell(j7 + 1, 3).value

    for j8 in range(line8sheet.max_row):
        if sheet.cell(i + 1, 1).value == line8sheet.cell(j8 + 1, 2).value:
            sheet.cell(i + 1, 2).value = line8sheet.cell(j8 + 1, 3).value

    for j11 in range(line11sheet.max_row):
        if sheet.cell(i + 1, 1).value == line11sheet.cell(j11 + 1, 2).value:
            sheet.cell(i + 1, 2).value = line11sheet.cell(j11 + 1, 3).value

    for j12 in range(line12sheet.max_row):
        if sheet.cell(i + 1, 1).value == line12sheet.cell(j12 + 1, 2).value:
            sheet.cell(i + 1, 2).value = line12sheet.cell(j12 + 1, 3).value

wb.save(filePath)

