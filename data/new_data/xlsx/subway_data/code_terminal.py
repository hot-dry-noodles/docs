# -*- coding: utf-8 -*-

import openpyxl

filePath = 'coding_terminal.xlsx'
wb = openpyxl.load_workbook(filePath)
sheet = wb['sheet1']

dfilePath = 'line12.xlsx'
dwb = openpyxl.load_workbook(dfilePath)
dsheet = dwb['sheet1']

count = 0

for i in range(dsheet.max_row):
	for j in range(sheet.max_row):
		if dsheet.cell(row = i + 1, column = 3).value == sheet.cell(row = j + 1, column = 2).value:
			dsheet.cell(row = i + 1, column = 2, value = sheet.cell(row = j + 1, column = 1).value)
			count += 1

print(count)

dwb.save(dfilePath)
