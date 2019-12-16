import openpyxl

begin_row = 2

filePath = r'data\new_data\xlsx\subway_data\line12.xlsx'
wb = openpyxl.load_workbook(filePath)
sheet = wb['sheet1']

for i in range(sheet.max_row - 1):
    sheet.cell(begin_row + i , 1).value += 1

wb.save(filePath)
print('finished!')