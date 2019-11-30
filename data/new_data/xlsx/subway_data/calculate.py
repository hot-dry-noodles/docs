# -*- coding: utf-8 -*-
# NOTICE:
# 	1.	operations for each column is excuted seperatedly but in order
# 	2.	efficiency is not taken in to considertaion for the following reasons: 
#  			- improve the readability
#  			- it is not related to the efficiency of the application
#  			- the efficency of the application is only related to the operations in database 

import openpyxl
import math

terminal_num = 198
row_count = 0

filePath = 'distance_route_price.xlsx'
wb = openpyxl.load_workbook(filePath)
sheet = wb['sheet1']

# line1Path = 'line1.xlsx'
# line1wb = openpyxl.load_workbook(line1Path)
# line1sheet = line1wb['sheet1']
#
# line2Path = 'line2.xlsx'
# line2wb = openpyxl.load_workbook(line2Path)
# line2sheet = line2wb['sheet1']
#
# line3Path = 'line3.xlsx'
# line3wb = openpyxl.load_workbook(line3Path)
# line3sheet = line3wb['sheet1']
#
# line4Path = 'line4.xlsx'
# line4wb = openpyxl.load_workbook(line4Path)
# line4sheet = line4wb['sheet1']
#
# line6Path = 'line6.xlsx'
# line6wb = openpyxl.load_workbook(line6Path)
# line6sheet = line6wb['sheet1']
#
# line7Path = 'line7.xlsx'
# line7wb = openpyxl.load_workbook(line7Path)
# line7sheet = line7wb['sheet1']
#
# line8Path = 'line8.xlsx'
# line8wb = openpyxl.load_workbook(line8Path)
# line8sheet = line8wb['sheet1']
#
# line11Path = 'line11.xlsx'
# line11wb = openpyxl.load_workbook(line11Path)
# line11sheet = line11wb['sheet1']
#
# line12Path = 'line12.xlsx'
# line12wb = openpyxl.load_workbook(line12Path)
# line12sheet = line12wb['sheet1']

codingTerminalPath = 'coding_terminal.xlsx'
codingTerminalwb = openpyxl.load_workbook(codingTerminalPath)
codingTerminalsheet = codingTerminalwb['sheet1']


def createList(sheet, column_num):
	""" create the list according to the column_num """
	clist = []
	for i in range(sheet.max_row):
		clist.append(sheet.cell(i + 1, column_num).value)
	return clist


# # lists of terminals in a line
# line1list = createList(line1sheet, 2)
# line2list = createList(line2sheet, 2)
# line3list = createList(line3sheet, 2)
# line4list = createList(line4sheet, 2)
# line6list = createList(line6sheet, 2)
# line7list = createList(line7sheet, 2)
# line8list = createList(line8sheet, 2)
# line11list = createList(line11sheet, 2)
# line12list = createList(line12sheet, 2)
#
# # lists of distances in a line
# dline1list = createList(line1sheet, 8)
# dline2list = createList(line2sheet, 8)
# dline3list = createList(line3sheet, 8)
# dline4list = createList(line4sheet, 8)
# dline6list = createList(line6sheet, 8)
# dline7list = createList(line7sheet, 8)
# dline8list = createList(line8sheet, 8)
# dline11list = createList(line11sheet, 8)
# dline12list = createList(line12sheet, 8)

# # column 1 - start terminal code
# for i in range(terminal_num * terminal_num):
# 	sheet.cell(i + 1, 1).value = i // terminal_num
#
# # column 2 - start terminal name
# for i in range(terminal_num * terminal_num):
#
# 	for j1 in range(line1sheet.max_row):
# 		if sheet.cell(i + 1, 1).value == line1sheet.cell(j1 + 1, 2).value:
# 			sheet.cell(i + 1, 2).value = line1sheet.cell(j1 + 1, 3).value
#
# 	for j2 in range(line2sheet.max_row):
# 		if sheet.cell(i + 1, 1).value == line2sheet.cell(j2 + 1, 2).value:
# 			sheet.cell(i + 1, 2).value = line2sheet.cell(j2 + 1, 3).value
#
# 	for j3 in range(line3sheet.max_row):
# 		if sheet.cell(i + 1, 1).value == line3sheet.cell(j3 + 1, 2).value:
# 			sheet.cell(i + 1, 2).value = line3sheet.cell(j3 + 1, 3).value
#
# 	for j4 in range(line4sheet.max_row):
# 		if sheet.cell(i + 1, 1).value == line4sheet.cell(j4 + 1, 2).value:
# 			sheet.cell(i + 1, 2).value = line4sheet.cell(j4 + 1, 3).value
#
# 	for j6 in range(line6sheet.max_row):
# 		if sheet.cell(i + 1, 1).value == line6sheet.cell(j6 + 1, 2).value:
# 			sheet.cell(i + 1, 2).value = line6sheet.cell(j6 + 1, 3).value
#
# 	for j7 in range(line7sheet.max_row):
# 		if sheet.cell(i + 1, 1).value == line7sheet.cell(j7 + 1, 2).value:
# 			sheet.cell(i + 1, 2).value = line7sheet.cell(j7 + 1, 3).value
#
# 	for j8 in range(line8sheet.max_row):
# 		if sheet.cell(i + 1, 1).value == line8sheet.cell(j8 + 1, 2).value:
# 			sheet.cell(i + 1, 2).value = line8sheet.cell(j8 + 1, 3).value
#
# 	for j11 in range(line11sheet.max_row):
# 		if sheet.cell(i + 1, 1).value == line11sheet.cell(j11 + 1, 2).value:
# 			sheet.cell(i + 1, 2).value = line11sheet.cell(j11 + 1, 3).value
#
# 	for j12 in range(line12sheet.max_row):
# 		if sheet.cell(i + 1, 1).value == line12sheet.cell(j12 + 1, 2).value:
# 			sheet.cell(i + 1, 2).value = line12sheet.cell(j12 + 1, 3).value
#
# # column 3 - end terminal code
# for i in range(terminal_num):
# 	for j in range(terminal_num):
# 		row_count += 1
# 		sheet.cell(row_count, 3).value = j
# row_count = 0
#
# # column 4 - end terminal name
# for i in range(terminal_num * terminal_num):
#
# 	for j1 in range(line1sheet.max_row):
# 		if sheet.cell(i + 1, 3).value == line1sheet.cell(j1 + 1, 2).value:
# 			sheet.cell(i + 1, 4).value = line1sheet.cell(j1 + 1, 3).value
#
# 	for j2 in range(line2sheet.max_row):
# 		if sheet.cell(i + 1, 3).value == line2sheet.cell(j2 + 1, 2).value:
# 			sheet.cell(i + 1, 4).value = line2sheet.cell(j2 + 1, 3).value
#
# 	for j3 in range(line3sheet.max_row):
# 		if sheet.cell(i + 1, 3).value == line3sheet.cell(j3 + 1, 2).value:
# 			sheet.cell(i + 1, 4).value = line3sheet.cell(j3 + 1, 3).value
#
# 	for j4 in range(line4sheet.max_row):
# 		if sheet.cell(i + 1, 3).value == line4sheet.cell(j4 + 1, 2).value:
# 			sheet.cell(i + 1, 4).value = line4sheet.cell(j4 + 1, 3).value
#
# 	for j6 in range(line6sheet.max_row):
# 		if sheet.cell(i + 1, 3).value == line6sheet.cell(j6 + 1, 2).value:
# 			sheet.cell(i + 1, 4).value = line6sheet.cell(j6 + 1, 3).value
#
# 	for j7 in range(line7sheet.max_row):
# 		if sheet.cell(i + 1, 3).value == line7sheet.cell(j7 + 1, 2).value:
# 			sheet.cell(i + 1, 4).value = line7sheet.cell(j7 + 1, 3).value
#
# 	for j8 in range(line8sheet.max_row):
# 		if sheet.cell(i + 1, 3).value == line8sheet.cell(j8 + 1, 2).value:
# 			sheet.cell(i + 1, 4).value = line8sheet.cell(j8 + 1, 3).value
#
# 	for j11 in range(line11sheet.max_row):
# 		if sheet.cell(i + 1, 3).value == line11sheet.cell(j11 + 1, 2).value:
# 			sheet.cell(i + 1, 4).value = line11sheet.cell(j11 + 1, 3).value
#
# 	for j12 in range(line12sheet.max_row):
# 		if sheet.cell(i + 1, 3).value == line12sheet.cell(j12 + 1, 2).value:
# 			sheet.cell(i + 1, 4).value = line12sheet.cell(j12 + 1, 3).value
#
# # column 5 - distance
#
# # 5.1 one-line distance
# for i in range(terminal_num * terminal_num):
#
# 	if sheet.cell(i + 1, 1).value in line1list and sheet.cell(i + 1, 3).value in line1list:
# 		index1 = line1list.index(sheet.cell(i + 1, 1).value)
# 		index2 = line1list.index(sheet.cell(i + 1, 3).value)
# 		if index1 <= index2:
# 			sheet.cell(i + 1, 5).value = sum(dline1list[index1 + 1 : index2 + 1])
# 		else:
# 			sheet.cell(i + 1, 5).value = sum(dline1list[index2 + 1 : index1 + 1])
#
# 	elif sheet.cell(i + 1, 1).value in line2list and sheet.cell(i + 1, 3).value in line2list:
# 		index1 = line2list.index(sheet.cell(i + 1, 1).value)
# 		index2 = line2list.index(sheet.cell(i + 1, 3).value)
# 		if index1 <= index2:
# 			sheet.cell(i + 1, 5).value = sum(dline2list[index1 + 1 : index2 + 1])
# 		else:
# 			sheet.cell(i + 1, 5).value = sum(dline2list[index2 + 1 : index1 + 1])
#
# 	elif sheet.cell(i + 1, 1).value in line3list and sheet.cell(i + 1, 3).value in line3list:
# 		index1 = line3list.index(sheet.cell(i + 1, 1).value)
# 		index2 = line3list.index(sheet.cell(i + 1, 3).value)
# 		if index1 <= index2:
# 			sheet.cell(i + 1, 5).value = sum(dline3list[index1 + 1 : index2 + 1])
# 		else:
# 			sheet.cell(i + 1, 5).value = sum(dline3list[index2 + 1 : index1 + 1])
#
# 	elif sheet.cell(i + 1, 1).value in line4list and sheet.cell(i + 1, 3).value in line4list:
# 		index1 = line4list.index(sheet.cell(i + 1, 1).value)
# 		index2 = line4list.index(sheet.cell(i + 1, 3).value)
# 		if index1 <= index2:
# 			sheet.cell(i + 1, 5).value = sum(dline4list[index1 + 1 : index2 + 1])
# 		else:
# 			sheet.cell(i + 1, 5).value = sum(dline4list[index2 + 1 : index1 + 1])
#
# 	elif sheet.cell(i + 1, 1).value in line6list and sheet.cell(i + 1, 3).value in line6list:
# 		index1 = line6list.index(sheet.cell(i + 1, 1).value)
# 		index2 = line6list.index(sheet.cell(i + 1, 3).value)
# 		if index1 <= index2:
# 			sheet.cell(i + 1, 5).value = sum(dline6list[index1 + 1 : index2 + 1])
# 		else:
# 			sheet.cell(i + 1, 5).value = sum(dline6list[index2 + 1 : index1 + 1])
#
# 	elif sheet.cell(i + 1, 1).value in line7list and sheet.cell(i + 1, 3).value in line7list:
# 		index1 = line7list.index(sheet.cell(i + 1, 1).value)
# 		index2 = line7list.index(sheet.cell(i + 1, 3).value)
# 		if index1 <= index2:
# 			sheet.cell(i + 1, 5).value = sum(dline7list[index1 + 1 : index2 + 1])
# 		else:
# 			sheet.cell(i + 1, 5).value = sum(dline7list[index2 + 1 : index1 + 1])
#
# 	elif sheet.cell(i + 1, 1).value in line8list and sheet.cell(i + 1, 3).value in line8list:
# 		index1 = line8list.index(sheet.cell(i + 1, 1).value)
# 		index2 = line8list.index(sheet.cell(i + 1, 3).value)
# 		if index1 <= index2:
# 			sheet.cell(i + 1, 5).value = sum(dline8list[index1 + 1 : index2 + 1])
# 		else:
# 			sheet.cell(i + 1, 5).value = sum(dline8list[index2 + 1 : index1 + 1])
#
# 	elif sheet.cell(i + 1, 1).value in line11list and sheet.cell(i + 1, 3).value in line11list:
# 		index1 = line11list.index(sheet.cell(i + 1, 1).value)
# 		index2 = line11list.index(sheet.cell(i + 1, 3).value)
# 		if index1 <= index2:
# 			sheet.cell(i + 1, 5).value = sum(dline11list[index1 + 1 : index2 + 1])
# 		else:
# 			sheet.cell(i + 1, 5).value = sum(dline11list[index2 + 1 : index1 + 1])
#
# 	elif sheet.cell(i + 1, 1).value in line12list and sheet.cell(i + 1, 3).value in line12list:
# 		index1 = line12list.index(sheet.cell(i + 1, 1).value)
# 		index2 = line12list.index(sheet.cell(i + 1, 3).value)
# 		if index1 <= index2:
# 			sheet.cell(i + 1, 5).value = sum(dline12list[index1 + 1 : index2 + 1])
# 		else:
# 			sheet.cell(i + 1, 5).value = sum(dline12list[index2 + 1 : index1 + 1])
#
# 	else:
# 		# set unreachable
# 		sheet.cell(i + 1, 5).value = 999999

# 5.2 cross-line distance

# dijkstra
# def cLength(x, y):
# 	for i in range(terminal_num * terminal_num):
# 		if sheet.cell(i + 1, 1).value == x and sheet.cell(i + 1, 3).value == y:
# 			return sheet.cell(i + 1, 5).value


# v_list = createList(codingTerminalsheet, 1)
# print('vlist:', v_list)

# start_num = 168
# end_num = 169

# for i in range(start_num, terminal_num):

# 	print('**********************************************')
# 	print('ternimal-', i)

# 	x_list = list()
# 	y_list = v_list.copy()
# 	d_list = [999999] * terminal_num

# 	x_list.append(i)
# 	y_list.remove(i)
# 	d_list[i] = 0

# 	print('xlist:', x_list)
# 	print('ylist:', y_list)
# 	print('vlist:', v_list)

# 	for y in iter(y_list):
# 		if cLength(i, y) != 999999:
# 			d_list[y] = cLength(i, y)

# 	while y_list:

# 		print('left terminal num:', len(y_list))

# 		min_distance = 999999
# 		for ty in iter(y_list):
# 			if d_list[ty] < min_distance:
# 				min_distance = d_list[ty]
# 				y = ty

# 		print(y)

# 		x_list.append(y)
# 		y_list.remove(y)

# 		print('xlist:', x_list)
# 		print('ylist:', y_list)

# 		for w in iter(y_list):
# 			if d_list[y] + cLength(y, w) < d_list[w]:
# 				d_list[w] = d_list[y] + cLength(y, w)

# 	print('d_list of terminal-', i)

# 	print('start filling the d_list of terminal-', i)

# 	for ir in range(terminal_num * terminal_num):
# 		start = sheet.cell(ir + 1, 1).value
# 		end = sheet.cell(ir + 1, 3).value
# 		if start == i:
# 			if sheet.cell(ir + 1, 5).value > d_list[end]:
# 				sheet.cell(ir + 1, 5).value = d_list[end]

# 	wb.save(filePath)
# 	print('finish filling the d_list of terminal-', i, ':', d_list)

# 	# debug
# 	if i == end_num:
# 		break

# column 6 - route

# 6.1 one-line route


# 6.2 cross-line route


# column 7 - price

def calPrice(distance):
	if distance == 0:
		return 0
	elif 0 < distance <= 4:
		return 2	# max: 2
	elif 4 < distance <= 12:
		return 2 + math.ceil((distance-4)/4)	# max: 4 
	elif 12 < distance <= 24:
		return 4 + math.ceil((distance-12)/6)	# max: 6
	elif 24 < distance <= 40:
		return 6 + math.ceil((distance-24)/8)	# max: 8 
	elif 40 < distance <= 50:
		return 8 + math.ceil((distance-40)/10) # max: 9r
	elif distance > 50:
		return 9 + math.ceil((distance-50)/20)

for i in range(terminal_num * terminal_num):
	distance = sheet.cell(i + 1, 5).value
	sheet.cell(i + 1, 7).value = calPrice(distance)

print('finished!')
wb.save(filePath)
