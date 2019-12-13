# -*- coding: utf-8 -*-

import openpyxl
import networkx as nx
import datetime

terminal_num = 198
row_count = 0

filePath = r'data\new_data\xlsx\subway_data\distance_route_price.xlsx'
wb = openpyxl.load_workbook(filePath)
sheet = wb['sheet1']

line1Path = r'data\new_data\xlsx\subway_data\line1.xlsx'
line1wb = openpyxl.load_workbook(line1Path)
line1sheet = line1wb['sheet1']

line2Path = r'data\new_data\xlsx\subway_data\line2.xlsx'
line2wb = openpyxl.load_workbook(line2Path)
line2sheet = line2wb['sheet1']

line3Path = r'data\new_data\xlsx\subway_data\line3.xlsx'
line3wb = openpyxl.load_workbook(line3Path)
line3sheet = line3wb['sheet1']

line4Path = r'data\new_data\xlsx\subway_data\line4.xlsx'
line4wb = openpyxl.load_workbook(line4Path)
line4sheet = line4wb['sheet1']

line6Path = r'data\new_data\xlsx\subway_data\line6.xlsx'
line6wb = openpyxl.load_workbook(line6Path)
line6sheet = line6wb['sheet1']

line7Path = r'data\new_data\xlsx\subway_data\line7.xlsx'
line7wb = openpyxl.load_workbook(line7Path)
line7sheet = line7wb['sheet1']

line8Path = r'data\new_data\xlsx\subway_data\line8.xlsx'
line8wb = openpyxl.load_workbook(line8Path)
line8sheet = line8wb['sheet1']

line11Path = r'data\new_data\xlsx\subway_data\line11.xlsx'
line11wb = openpyxl.load_workbook(line11Path)
line11sheet = line11wb['sheet1']

line12Path = r'data\new_data\xlsx\subway_data\line12.xlsx'
line12wb = openpyxl.load_workbook(line12Path)
line12sheet = line12wb['sheet1']

def addEdge(G, _sheet):
    """ add all edges of a sheet """
    for i in range(_sheet.max_row - 2):
        start = _sheet.cell(i + 2, 1).value
        end = _sheet.cell(i + 3, 1).value
        distance = _sheet.cell(i + 3, 9).value
        G.add_edge(start, end, weight=distance)
        G.add_edge(end, start, weight=distance)
    return G


def Dijkstra(G, start, end):
    RG = G.reverse()
    dist = {}
    previous = {}
    for v in RG.nodes():
        dist[v] = float('inf')
        previous[v] = 'none'
    dist[end] = 0
    u = end
    while u != start:
        u = min(dist, key=dist.get)
        distu = dist[u]
        del dist[u]
        for u, v in RG.edges(u):
            if v in dist:
                alt = distu + RG[u][v]['weight']
                if alt < dist[v]:
                    dist[v] = alt
                    previous[v] = u
    path = (start,)
    last = start
    while last != end:
        nxt = previous[last]
        path += (nxt,)
        last = nxt
    return path


G = nx.DiGraph()
G = addEdge(G, line1sheet)
G = addEdge(G, line2sheet)
G = addEdge(G, line3sheet)
G = addEdge(G, line4sheet)
G = addEdge(G, line6sheet)
G = addEdge(G, line7sheet)
G = addEdge(G, line8sheet)
G = addEdge(G, line11sheet)
G = addEdge(G, line12sheet)


def createList(sheet, column_num):
	""" create the list according to the column_num """
	clist = list()
	for i in range(sheet.max_row - 1):
		clist.append(sheet.cell(i + 2, column_num).value)
	return clist

# lists of terminals in a line
line1list = createList(line1sheet, 1)
line2list = createList(line2sheet, 1)
line3list = createList(line3sheet, 1)
line4list = createList(line4sheet, 1)
line6list = createList(line6sheet, 1)
line7list = createList(line7sheet, 1)
line8list = createList(line8sheet, 1)
line11list = createList(line11sheet, 1)
line12list = createList(line12sheet, 1)

def minNums(startTime, endTime):
    """ calculate the minutes between 2 datetime """
    
    startTime1 = str(startTime)
    endTime1 = str(endTime)
    try:
        temp = datetime.datetime.strptime(startTime1, "%Y-%m-%d %H:%M:%S")
        startTime2 = datetime.datetime.strptime(startTime1[-8:], "%H:%M:%S")
    except:
        startTime2 = datetime.datetime.strptime(startTime1, "%H:%M:%S")
    try:
        temp = datetime.datetime.strptime(endTime1, "%Y-%m-%d %H:%M:%S")
        endTime2 = datetime.datetime.strptime(endTime1[-8:], "%H:%M:%S")
    except:
        endTime2 = datetime.datetime.strptime(endTime1, "%H:%M:%S")    
    seconds = (endTime2 - startTime2).seconds
    total_seconds = (endTime2 - startTime2).total_seconds()   
    mins = total_seconds / 60
    if int(mins) < 0:
        mins += 1440
    return int(mins)

def singleTime(start, end):
    """ return the time from begin to end in a line """
    sgtime = 0

    if start in line1list and end in line1list:
        i = line1list.index(start) + 2
        j = line1list.index(end) + 2
        stime = line1sheet.cell(i, 6).value
        etime = line1sheet.cell(j, 6).value
        sgtime = minNums(stime, etime)
    
    elif start in line2list and end in line2list:
        i = line2list.index(start) + 2
        j = line2list.index(end) + 2
        stime = line2sheet.cell(i, 6).value
        etime = line2sheet.cell(j, 6).value
        sgtime = minNums(stime, etime)       

    elif start in line3list and end in line3list:
        i = line3list.index(start) + 2
        j = line3list.index(end) + 2
        stime = line3sheet.cell(i, 6).value
        etime = line3sheet.cell(j, 6).value
        sgtime = minNums(stime, etime)   

    elif start in line4list and end in line4list:
        i = line4list.index(start) + 2
        j = line4list.index(end) + 2
        stime = line4sheet.cell(i, 6).value
        etime = line4sheet.cell(j, 6).value
        sgtime = minNums(stime, etime) 

    elif start in line6list and end in line6list:
        i = line6list.index(start) + 2
        j = line6list.index(end) + 2
        stime = line6sheet.cell(i, 6).value
        etime = line6sheet.cell(j, 6).value
        sgtime = minNums(stime, etime) 

    elif start in line7list and end in line7list:
        i = line7list.index(start) + 2
        j = line7list.index(end) + 2
        stime = line7sheet.cell(i, 6).value
        etime = line7sheet.cell(j, 6).value
        sgtime = minNums(stime, etime) 

    elif start in line8list and end in line8list:
        i = line8list.index(start) + 2
        j = line8list.index(end) + 2
        stime = line8sheet.cell(i, 6).value
        etime = line8sheet.cell(j, 6).value
        sgtime = minNums(stime, etime) 

    elif start in line11list and end in line11list:
        i = line11list.index(start) + 2
        j = line11list.index(end) + 2
        stime = line11sheet.cell(i, 6).value
        etime = line11sheet.cell(j, 6).value
        sgtime = minNums(stime, etime) 

    elif start in line12list and end in line12list:
        i = line12list.index(start) + 2
        j = line12list.index(end) + 2
        stime = line12sheet.cell(i, 6).value
        etime = line12sheet.cell(j, 6).value
        sgtime = minNums(stime, etime) 

    return sgtime


def getTime(start, end, route):
    """ return the time form start to end """
    dtime = 0
    tnum = len(route)
    if tnum == 1:
        dtime = 0
    for i in range(tnum - 1):
        dtime += singleTime(i, i + 1)
    return dtime

counter = 0
run = 0
for i in range(terminal_num * terminal_num):
    start = sheet.cell(i + 2, 1).value
    end = sheet.cell(i + 2, 3).value
    route = Dijkstra(G, start, end)

    # sheet.cell(i + 2, 6).value = str(route)
    # sheet.cell(i + 2, 7).value = len(route) - 1
  
    sheet.cell(i + 2, 9).value = str(getTime(start, end, route))

    counter += 1
    if counter == 198:
        print('run ', run, 'finished!')
        run += 1
        counter = 0

print('finished!')
wb.save(filePath)
