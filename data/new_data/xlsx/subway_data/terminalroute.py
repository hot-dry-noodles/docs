import re
import openpyxl

def dupe(items):
      seen = set()
      for item in items:
        if item not in seen:
          yield item
          seen.add(item)

def convertTerminal(s):

    if len(s) != 1:
        
        temp = re.findall(r'[0-9]+|[a-z]+', s)
        list_line = list()
        list_terminal = list()
        for i in iter(temp):
            if i.isdigit():
                list_terminal.append(i)
            else:
                list_line.append(i)
       
        list_line_new = list(dupe(list_line))
        list_terminal_new = list()
        for line in iter(list_line_new):
            list_terminal_new.append(list_terminal[list_line.index(line)])
        list_terminal_new.append(list_terminal[-1])

        finalList = []
        for i in range(len(list_line_new)):
            finalList.append(list_terminal_new[i])
            finalList.append(list_line_new[i])
        finalList.append(list_terminal_new[-1])

        _s =''.join(finalList)

        return _s
    
    else:
        return s

filePath = r'data\new_data\xlsx\subway_data\distance_route_price.xlsx'
wb = openpyxl.load_workbook(filePath)
sheet = wb['sheet1']
terminal_num = 198

for i in range(terminal_num * terminal_num):
    sheet.cell(i + 2, 6).value = convertTerminal(str(sheet.cell(i + 2, 6).value))

print('finished!')
wb.save(filePath)