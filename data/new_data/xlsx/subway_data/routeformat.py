import openpyxl
import copy

terminal_num = 198

filePath = r'data\new_data\xlsx\subway_data\distance_route_price.xlsx'
wb = openpyxl.load_workbook(filePath)
sheet = wb['sheet1']

line1Path = r'data\new_data\xlsx\subway_data\line1.xlsx'
line1wb = openpyxl.load_workbook(line1Path)
line1sheet = line1wb['sheet1']

line2Path = r'data\new_data\xlsx\subway_data\line2.xlsx'
line2wb = openpyxl.load_workbook(line2Path)
line2sheet = line2wb['sheet1']

line3Path = r'data\new_data\xlsx\subway_data\line3.xlsx'
line3wb = openpyxl.load_workbook(line3Path)
line3sheet = line3wb['sheet1']

line4Path = r'data\new_data\xlsx\subway_data\line4.xlsx'
line4wb = openpyxl.load_workbook(line4Path)
line4sheet = line4wb['sheet1']

line6Path = r'data\new_data\xlsx\subway_data\line6.xlsx'
line6wb = openpyxl.load_workbook(line6Path)
line6sheet = line6wb['sheet1']

line7Path = r'data\new_data\xlsx\subway_data\line7.xlsx'
line7wb = openpyxl.load_workbook(line7Path)
line7sheet = line7wb['sheet1']

line8Path = r'data\new_data\xlsx\subway_data\line8.xlsx'
line8wb = openpyxl.load_workbook(line8Path)
line8sheet = line8wb['sheet1']

line11Path = r'data\new_data\xlsx\subway_data\line11.xlsx'
line11wb = openpyxl.load_workbook(line11Path)
line11sheet = line11wb['sheet1']

line12Path = r'data\new_data\xlsx\subway_data\line12.xlsx'
line12wb = openpyxl.load_workbook(line12Path)
line12sheet = line12wb['sheet1']

def createList(sheet, column_num):
	""" create the list according to the column_num """
	clist = list()
	for i in range(sheet.max_row):
		clist.append(sheet.cell(i + 1, column_num).value)
	return clist

# lists of terminals in a line
line1list = createList(line1sheet, 2)
line2list = createList(line2sheet, 2)
line3list = createList(line3sheet, 2)
line4list = createList(line4sheet, 2)
line6list = createList(line6sheet, 2)
line7list = createList(line7sheet, 2)
line8list = createList(line8sheet, 2)
line11list = createList(line11sheet, 2)
line12list = createList(line12sheet, 2)

def sameLine(x, y):
	""" return 1 if x and y is in the same line """
	if x in line1list and y in line1list:
		return 'a'
	elif x in line2list and y in line2list:
		return 'b'
	elif x in line3list and y in line3list:
		return 'c'
	elif x in line4list and y in line4list:
		return 'd'
	elif x in line6list and y in line6list:
		return 'e'
	elif x in line7list and y in line7list:
		return 'f'
	elif x in line8list and y in line8list:
		return 'g'
	elif x in line11list and y in line11list:
		return 'h'
	elif x in line12list and y in line12list:
		return 'i'
	else:
		return '-'

for i in range(terminal_num * terminal_num):
	if sheet.cell(i + 1, 1).value == sheet.cell(i + 1, 3).value:
		continue
	else:
		token = str(sheet.cell(i + 1, 6).value).split('-')
		route = ''
		pre_nodeline = ''
		for w in range(len(token) - 1):
			
			node = int(token[w])
			next_node = int(token[w + 1])
			nodeline = sameLine(node, next_node)			
			route += str(node)
			route += nodeline

		route += str(token[-1])
	sheet.cell(i + 1, 6).value = str(route)

print('finished!')
wb.save(filePath)