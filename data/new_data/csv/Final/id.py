import openpyxl

terminal_num = 198
begin_row = 2
row_count = 0

filePath = r'data\new_data\csv\Final\booking_route.xlsx'
wb = openpyxl.load_workbook(filePath)
sheet = wb['Sheet1']

for i in range(terminal_num * terminal_num):
    sheet.cell(i + begin_row, 4).value = i // terminal_num + 1

for i in range(terminal_num):
	for j in range(terminal_num):
		sheet.cell(begin_row + row_count, 5).value = j + 1
		row_count += 1

wb.save(filePath)