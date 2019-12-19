#!/usr/bin/env python3
import json
from openpyxl import load_workbook

wb_name = 'line_points.xlsx'

wb = load_workbook(wb_name)
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    stations = {}
    for row in range(1, sheet.max_row):
        point = {}
        sid   = sheet['A{}'.format(row)].value
        sname = sheet['B{}'.format(row)].value
        lng   = sheet['C{}'.format(row)].value
        lat   = sheet['D{}'.format(row)].value
        point['name'] = sname
        point['lng'] = float(lng) * 1.5
        point['lat'] = float(800 - lat) * 1.5
        stations[sid] = point
    # with open(sheet_name + '.json', 'w', encoding='utf8') as out:
    #     json.dump(stations, out, ensure_ascii=False)
    with open(sheet_name + '.txt', 'w', encoding='utf8') as out:
        for p in stations:
            out.write(json.dumps(stations[p], ensure_ascii=False))
            out.write(',\n')
